import re
import os
import json
import logging
from typing import List, Dict, Optional, Tuple
from pathlib import Path
import openpyxl

# 配置日志
logger = logging.getLogger(__name__)

class CourseManager:
    """课程表管理系统，处理文本保存、解析和JSON导出"""
    
    def __init__(self, txt_path: str = "YongHuKeBiao.txt", json_path: str = "data.json", username: str = None):
        """
        初始化课程管理器
        
        Args:
            txt_path: 课表文本文件路径
            json_path: 输出JSON文件路径
            username: 用户名，如果提供则保存到用户特定文件
        """
        self.txt_path = Path(txt_path)
        self.username = username
        if username:
            # 如果提供了用户名，使用用户特定的JSON文件
            self.json_path = Path(f"user_{username}_schedule.json")
        else:
            # 否则使用默认的全局JSON文件
            self.json_path = Path(json_path)
        self.weekday_map = {
            '星期一': 1, '星期二': 2, '星期三': 3, '星期四': 4, 
            '星期五': 5, '星期六': 6, '星期日': 7
        }
        self.weekday_text_map = {
            1: '星期一', 2: '星期二', 3: '星期三', 
            4: '星期四', 5: '星期五', 6: '星期六', 7: '星期日'
        }
        logger.info(f"CourseManager initialized with txt_path: {txt_path}, json_path: {self.json_path}, username: {username}")
    
    def save_text_to_file(self, text: str) -> None:
        """
        将用户文本保存到TXT文件
        
        Args:
            text: 要保存的文本内容
            
        Raises:
            IOError: 文件写入失败时抛出
        """
        try:
            self.txt_path.parent.mkdir(parents=True, exist_ok=True)
            with open(self.txt_path, 'w', encoding='utf-8') as f:
                f.write(text)
            logger.info(f"Successfully saved text to {self.txt_path}")
        except Exception as e:
            logger.error(f"Failed to save text to {self.txt_path}: {e}")
            raise IOError(f"保存文件失败: {e}")
    
    def _parse_weekday_line(self, line: str) -> Optional[int]:
        """
        解析星期行
        
        Args:
            line: 待解析的行
            
        Returns:
            星期数字(1-7)或None
        """
        weekday_match = re.match(r'^(星期[一二三四五六日])$', line)
        if weekday_match:
            weekday_text = weekday_match.group(1)
            return self.weekday_map.get(weekday_text)
        return None
    
    def _parse_time_week_line(self, line: str) -> Optional[Tuple[int, int, int, int]]:
        """
        解析节次和周次行
        
        Args:
            line: 待解析的行
            
        Returns:
            (开始节次, 结束节次, 开始周次, 结束周次)或None
        """
        time_week_match = re.match(r'^(\d+)-(\d+)节 \((\d+)-(\d+)周\)$', line)
        if time_week_match:
            return (
                int(time_week_match.group(1)),  # start_section
                int(time_week_match.group(2)),  # end_section
                int(time_week_match.group(3)),  # start_week
                int(time_week_match.group(4))   # end_week
            )
        return None
    
    def _create_course_dict(self, course_name: str, location: str, 
                           current_weekday: int, start_section: int, 
                           end_section: int, start_week: int, end_week: int) -> Dict:
        """
        创建课程字典
        
        Args:
            course_name: 课程名称
            location: 教室位置
            current_weekday: 星期几
            start_section: 开始节次
            end_section: 结束节次
            start_week: 开始周次
            end_week: 结束周次
            
        Returns:
            课程字典
        """
        return {
            'course_name': course_name,
            'location': location,
            'day_of_week': current_weekday,
            'start_section': start_section,
            'end_section': end_section,
            'weeks': f"{start_week}-{end_week}"
        }
    
    def parse_course_text(self) -> List[Dict]:
        """
        从TXT文件解析课程数据，支持无空行分隔的课表格式
        
        Returns:
            课程列表
            
        Raises:
            FileNotFoundError: 文件不存在时抛出
            ValueError: 解析失败时抛出
        """
        if not self.txt_path.exists():
            error_msg = f"找不到文件: {self.txt_path}"
            logger.error(error_msg)
            raise FileNotFoundError(error_msg)
        
        try:
            with open(self.txt_path, 'r', encoding='utf-8') as f:
                text = f.read()
            
            courses = []
            current_weekday = None
            lines = [line.strip() for line in text.split('\n') if line.strip()]
            
            logger.info(f"开始解析课表，共 {len(lines)} 行数据")
            
            i = 0
            while i < len(lines):
                line = lines[i]
                
                # 检查是否为星期行
                weekday = self._parse_weekday_line(line)
                if weekday is not None:
                    current_weekday = weekday
                    logger.debug(f"解析到星期: {self.weekday_text_map[current_weekday]}")
                    i += 1
                    continue
                
                # 如果当前没有有效的星期，跳过后续解析
                if current_weekday is None:
                    i += 1
                    continue
                
                # 检查是否为节次和周次行
                time_week_data = self._parse_time_week_line(line)
                if time_week_data:
                    start_section, end_section, start_week, end_week = time_week_data
                    
                    # 验证数据有效性
                    if start_section > end_section or start_week > end_week:
                        logger.warning(f"无效的节次或周次范围: {line}")
                        i += 1
                        continue
                    
                    # 获取课程名称和教室
                    if i + 2 >= len(lines):
                        logger.warning(f"课程信息不完整，跳过: {line}")
                        i += 1
                        continue
                    
                    course_name = lines[i + 1].strip()
                    location = lines[i + 2].strip()
                    
                    if not course_name or not location:
                        logger.warning(f"课程名称或教室为空，跳过: {line}")
                        i += 1
                        continue
                    
                    # 创建课程数据
                    course = self._create_course_dict(
                        course_name, location, current_weekday,
                        start_section, end_section, start_week, end_week
                    )
                    courses.append(course)
                    
                    logger.debug(f"解析到课程: {course_name} - {location}")
                    i += 3
                    continue
                
                i += 1
            
            logger.info(f"课表解析完成，共解析到 {len(courses)} 门课程")
            return courses
            
        except Exception as e:
            error_msg = f"解析课表文件失败: {e}"
            logger.error(error_msg)
            raise ValueError(error_msg)
    
    def convert_to_json(self, courses: List[Dict]) -> Dict:
        """
        将课程列表转换为结构化JSON数据，保证每天课程按节次排序
        
        Args:
            courses: 课程列表
            
        Returns:
            按星期分组的课程数据
        """
        # 按星期分组
        grouped_data = {i: [] for i in range(1, 8)}  # 1-7对应周一到周日
        
        for course in courses:
            day = course['day_of_week']
            grouped_data[day].append(course)
        
        # 每天按 start_section 排序
        for day in grouped_data:
            grouped_data[day].sort(key=lambda x: x['start_section'])
        
        # 转换为星期文本格式
        result = {
            self.weekday_text_map[day]: [
                {
                    'course_name': c['course_name'],
                    'location': c['location'],
                    'time': f"{c['start_section']}-{c['end_section']}节",
                    'weeks': c['weeks']
                } for c in grouped_data[day]
            ] for day in grouped_data if grouped_data[day]
        }
        
        logger.info(f"JSON转换完成，共 {len(result)} 天有课程")
        return result
    
    def save_to_json(self, data: Dict) -> None:
        """
        将数据保存到JSON文件
        
        Args:
            data: 要保存的数据
            
        Raises:
            IOError: 文件写入失败时抛出
        """
        try:
            self.json_path.parent.mkdir(parents=True, exist_ok=True)
            with open(self.json_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=4)
            logger.info(f"Successfully saved JSON to {self.json_path}")
        except Exception as e:
            error_msg = f"保存JSON文件失败: {e}"
            logger.error(error_msg)
            raise IOError(error_msg)
    
    def process_user_text(self, text: str) -> Dict:
        """
        处理用户文本的完整流程
        
        Args:
            text: 用户输入的课表文本
            
        Returns:
            处理后的JSON数据
            
        Raises:
            ValueError: 文本处理失败时抛出
            IOError: 文件操作失败时抛出
        """
        if not text or not text.strip():
            error_msg = "输入的课表文本为空"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        logger.info("开始处理用户课表文本")
        
        try:
            # 保存文本到TXT
            self.save_text_to_file(text)
            
            # 解析课程数据
            courses = self.parse_course_text()
            
            # 转换为JSON格式
            json_data = self.convert_to_json(courses)
            
            # 保存到JSON文件
            self.save_to_json(json_data)
            
            logger.info("用户课表文本处理完成")
            return json_data
            
        except Exception as e:
            error_msg = f"处理用户文本失败: {e}"
            logger.error(error_msg)
            raise ValueError(error_msg)
    
    def process_user_excel(self, file_path: str) -> dict:
        """
        解析截图格式的 Excel 课表，整理为 data.json 所需结构。
        Args:
            file_path: Excel 文件路径
        Returns:
            处理后的JSON数据
        """
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # 第一行为星期，第一列为节次
        days = [cell.value for cell in ws[1][1:]]  # 跳过A1
        schedule = {day: [] for day in days if day}
        for row in ws.iter_rows(min_row=2):
            section = row[0].value  # 节次（如1,2,3...）
            for col, cell in enumerate(row[1:], start=1):
                day = days[col-1]
                if not day or not cell.value:
                    continue
                # 支持多门课用换行分隔
                for entry in str(cell.value).split('\n'):
                    parts = entry.split(',')
                    if len(parts) < 3:
                        continue
                    time_part = parts[0].strip()  # 3-4节 (1-17周)
                    course_name = parts[1].strip()
                    location = parts[2].strip()
                    # 解析节次和周次
                    import re
                    m = re.match(r'(\d+)-(\d+)节\s*\((\d+)-(\d+)周\)', time_part)
                    if m:
                        start_section, end_section, start_week, end_week = m.groups()
                        course_obj = {
                            'course_name': course_name,
                            'location': location,
                            'time': f"{start_section}-{end_section}节",
                            'weeks': f"{start_week}-{end_week}"
                        }
                        # 去重：只添加未出现过的课程
                        if course_obj not in schedule[day]:
                            schedule[day].append(course_obj)
        # 保存到json
        self.save_to_json(schedule)
        return schedule
    
    def get_course_count(self) -> Dict[str, int]:
        """
        获取每天的课程数量统计
        
        Returns:
            每天课程数量的字典
        """
        try:
            courses = self.parse_course_text()
            grouped_data = {i: [] for i in range(1, 8)}
            
            for course in courses:
                day = course['day_of_week']
                grouped_data[day].append(course)
            
            return {
                self.weekday_text_map[day]: len(courses)
                for day in grouped_data if grouped_data[day]
            }
        except Exception as e:
            logger.error(f"获取课程统计失败: {e}")
            return {}

